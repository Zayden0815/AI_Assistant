from pathlib import Path
from datetime import datetime
import json
import re
import html

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
REPORT_DIR = BASE_DIR / "reports"
REPORT_DIR.mkdir(exist_ok=True)

IMPORT_COLUMNS = [
    "tc_id", "title", "given", "when", "then",
    "expected_action", "verification_method", "methodology"
]


def safe_text(value):
    if value is None:
        return ""
    return str(value).strip()


def html_text(value):
    return html.escape(safe_text(value)).replace("\n", "<br/>")


def normalize_testcases(cases, default_method="ATDD"):
    normalized = []
    for i, tc in enumerate(cases or [], start=1):
        if not isinstance(tc, dict):
            continue
        normalized.append({
            "tc_id": safe_text(tc.get("tc_id") or tc.get("TC ID") or tc.get("TC_ID") or f"ATDD-{i:03d}"),
            "title": safe_text(tc.get("title") or tc.get("Title")),
            "given": safe_text(tc.get("given") or tc.get("Given") or tc.get("arrange") or tc.get("Arrange") or tc.get("acceptance_criteria") or tc.get("Acceptance Criteria")),
            "when": safe_text(tc.get("when") or tc.get("When") or tc.get("act") or tc.get("Act") or tc.get("trigger") or tc.get("Trigger")),
            "then": safe_text(tc.get("then") or tc.get("Then") or tc.get("assert") or tc.get("Assert") or tc.get("expected_result") or tc.get("Expected Result")),
            "expected_action": safe_text(tc.get("expected_action") or tc.get("Expected Action") or tc.get("expected") or tc.get("Expected")),
            "verification_method": safe_text(tc.get("verification_method") or tc.get("Verification Method") or tc.get("method") or tc.get("Method")),
            "methodology": safe_text(tc.get("methodology") or tc.get("Methodology") or tc.get("development_method") or tc.get("Development Method") or default_method),
        })
    return normalized


def parse_test_cases_from_text(text, default_method="ATDD"):
    text = safe_text(text)
    if not text:
        return []
    cases = []
    blocks = re.split(r"\n\s*\n", text)
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue
        first = lines[0]
        if " - " in first:
            tc_id, title = first.split(" - ", 1)
        else:
            tc_id, title = first, ""
        if not re.search(r"(TC|ATDD|BDD|TDD|DDD)[-_]?\d+", tc_id, re.I):
            continue

        def after(*prefixes):
            for prefix in prefixes:
                for line in lines:
                    if line.lower().startswith(prefix.lower()):
                        return line.split(":", 1)[1].strip() if ":" in line else line.replace(prefix, "", 1).strip()
            return ""

        cases.append({
            "tc_id": tc_id,
            "title": title,
            "given": after("Given:", "Arrange:", "Acceptance Criteria:", "Domain:"),
            "when": after("When:", "Act:", "Trigger:"),
            "then": after("Then:", "Assert:", "Expected Rule:"),
            "expected_action": after("Expected Action:", "Expected:"),
            "verification_method": after("Verification Method:", "Method:"),
            "methodology": after("Methodology:", "Development Method:") or default_method,
        })
    return normalize_testcases(cases, default_method)


def get_testcases(data):
    method = data.get("development_method", "ATDD")
    if isinstance(data.get("test_cases"), list) and data.get("test_cases"):
        return normalize_testcases(data.get("test_cases"), method)
    return parse_test_cases_from_text(data.get("testcase", ""), method)


def export_json(data):
    method = data.get("development_method", "ATDD")
    filename = f"ISO26262_{method}_ImportReady_{datetime.now():%Y%m%d_%H%M%S}.json"
    path = REPORT_DIR / filename

    payload = {
        "requirement": data.get("requirement", ""),
        "development_method": method,
        "related_parts": data.get("related_parts", ""),
        "referenced_pages": data.get("referenced_pages", ""),
        "asil_candidate": data.get("asil_candidate", ""),
        "asil_basis": data.get("asil_basis", ""),
        "test_cases": get_testcases(data),
        "traceability": data.get("traceability") or {},
    }

    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(path)


def export_excel(data):
    method = data.get("development_method", "ATDD")
    filename = f"ISO26262_{method}_ImportReady_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = REPORT_DIR / filename
    testcases = get_testcases(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.append(IMPORT_COLUMNS)
    for tc in testcases:
        ws.append([tc.get(col, "") for col in IMPORT_COLUMNS])
    style_sheet(ws)

    summary = wb.create_sheet("Summary")
    summary.append(["Field", "Value"])
    summary.append(["Document Type", "Streamlit Import Ready Test Case Sheet"])
    summary.append(["Development Method", method])
    summary.append(["Requirement", data.get("requirement", "")])
    summary.append(["Related ISO Parts", data.get("related_parts", "")])
    summary.append(["Referenced Pages", data.get("referenced_pages", "")])
    summary.append(["ASIL Candidate", data.get("asil_candidate", "")])
    summary.append(["ASIL Basis", data.get("asil_basis", "")])
    summary.append(["Total Test Cases", len(testcases)])
    style_sheet(summary)

    wb.save(path)
    return str(path)


def export_pdf(data):
    method = data.get("development_method", "ATDD")
    filename = f"ISO26262_{method}_ImportReady_{datetime.now():%Y%m%d_%H%M%S}.pdf"
    path = REPORT_DIR / filename
    testcases = get_testcases(data)

    doc = SimpleDocTemplate(str(path), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("<b>ISO26262 Functional Safety Test Case Report</b>", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Development Method: <b>{html_text(method)}</b>", styles["BodyText"]))
    story.append(Paragraph("Format: <b>Streamlit Import Compatible</b>", styles["BodyText"]))
    story.append(Paragraph(f"Requirement: {html_text(data.get('requirement', ''))}", styles["BodyText"]))
    story.append(Paragraph(f"ASIL Candidate: {html_text(data.get('asil_candidate', ''))}", styles["BodyText"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("<b>[[TEST_CASES_IMPORT_BLOCK]]</b>", styles["Code"]))
    for tc in testcases:
        block = "\n".join([
            f"TC_ID={tc.get('tc_id', '')}",
            f"TITLE={tc.get('title', '')}",
            f"GIVEN={tc.get('given', '')}",
            f"WHEN={tc.get('when', '')}",
            f"THEN={tc.get('then', '')}",
            f"EXPECTED_ACTION={tc.get('expected_action', '')}",
            f"VERIFICATION_METHOD={tc.get('verification_method', '')}",
            f"METHODOLOGY={tc.get('methodology', method)}",
            "---",
        ])
        story.append(Paragraph(html_text(block), styles["Code"]))
        story.append(Spacer(1, 4))
    story.append(Paragraph("<b>[[END_TEST_CASES_IMPORT_BLOCK]]</b>", styles["Code"]))

    story.append(PageBreak())
    story.append(Paragraph("<b>ISO26262 Evidence</b>", styles["Heading2"]))
    story.append(Paragraph(html_text(data.get("iso", "")), styles["BodyText"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("<b>Hazard / ASIL</b>", styles["Heading2"]))
    story.append(Paragraph(html_text(data.get("hazard", "")), styles["BodyText"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("<b>Safety Goal / FSR / TSR</b>", styles["Heading2"]))
    story.append(Paragraph(html_text(data.get("safety", "")), styles["BodyText"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph("<b>Traceability</b>", styles["Heading2"]))
    story.append(Paragraph(html_text(data.get("trace", "")), styles["BodyText"]))

    doc.build(story)
    return str(path)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 12
        for cell in col:
            if cell.value:
                max_len = min(max(max_len, len(str(cell.value)) + 2), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len